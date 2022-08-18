import os
from abc import ABC

from openpyxl import load_workbook

from .exceptions import ImproperlyConfigured
from .fields import BaseField


class BaseSheet(ABC):
    header_rows = 1
    sheet_name = None
    label_row = None

    def __init__(self, path=None, file=None):
        self.path = path
        self.file = file
        self._fields = None
        self._rows = []
        self._infos = []
        self._errors = []

        self.check()

    def check(self):
        if self.get_header_rows() < 1:
            raise ImproperlyConfigured("Sheets must have at least one header row")

        if self.path is None and self.file is None:
            raise ImproperlyConfigured("Either path or file has to be passed")

        if not self.sheet_name:
            raise ImproperlyConfigured("No sheet name set")

    def load(self):
        header_rows = self.get_header_rows()

        sheet = self.__get_sheet()

        if self.has_errors:
            return

        column_map = self.__column_map(sheet)
        self.__check_column_map(column_map)

        if self.has_errors:
            return

        for row_cells in sheet.iter_rows(min_row=header_rows + 1):
            print(row_cells)

    def get_sheet_name(self):
        return self.sheet_name

    def get_header_rows(self):
        return self.header_rows

    def get_label_row(self):
        if self.label_row:
            return self.label_row
        return self.get_header_rows() - 1

    def get_file(self):
        if not self.file:
            if os.path.exists(self.path):
                self.file = open(self.path, "r")
            else:
                raise ValueError("%s is not a file" % self.path)
        return self.file

    @property
    def errors(self):
        return self._errors

    @property
    def has_errors(self):
        return len(self._errors) > 0

    @property
    def infos(self):
        return self._infos

    @property
    def has_infos(self):
        return len(self._infos) > 0

    def __enter__(self):
        if not self.file:
            self.get_file()

    def __exit__(self, exc_type, exc_value, traceback):
        if self.file:
            self.file.close()

    def __getitem__(self, item):
        if item < 0 or (item > len(self) - 1):
            raise IndexError()

        return self._rows[item]

    def __len__(self):
        return len(self._rows)

    def _get_fields(self):
        if self._fields is None:
            self._fields = []
            for attr, value in self.__class__.__dict__.items():
                if isinstance(value, BaseField):
                    self._fields.append(value)

        return self._fields

    def _add_error(self, message, index=None):
        # Add to row index, if it’s related to a row.
        if isinstance(index, int):
            message = f"Row {index + 1}: {message}"
        self._errors.append(message)

    def _add_info(self, message, index=None):
        # Add to row index, if it’s related to a row.
        if isinstance(index, int):
            message = f"Row {index + 1}: {message}"
        self._infos.append(message)

    def __column_map(self, sheet):
        column_map = {}
        label_row = self.get_label_row()

        for index, column in enumerate(
            sheet.iter_cols(0, sheet.max_column, min_row=label_row, max_row=label_row)
        ):
            column_map[column[0].value] = index

        return column_map

    def __check_column_map(self, column_map):
        fields = self._get_fields()
        used_fields = set([field.source for field in fields])
        all_fields = set(column_map.keys())

        for field in used_fields:
            if field not in all_fields:
                self._add_error(f"Field {field} is missing in sheet")

    def __get_sheet(self):
        wb = load_workbook(filename=self.path)
        sheet_name = self.get_sheet_name()

        try:
            return wb[sheet_name]
        except KeyError:
            self._add_error(f"There’s no sheet {sheet_name} in spreadsheet {self.path}")
