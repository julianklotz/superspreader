from abc import ABC

from openpyxl import load_workbook

from .exceptions import ImproperlyConfigured, ValidationException
from .fields import BaseField
from .i18n import EN
from .i18n import translate as _


class BaseSheet(ABC):
    header_rows = 1
    sheet_name = None
    label_row = None

    def __init__(self, path, language=EN, extra_data=None):
        self.language = language
        self.path = path
        self._fields = self._build_fields()
        self._rows = []
        self._infos = []
        self._errors = []

        if extra_data is not None:
            assert isinstance(extra_data, dict)
            self._extra_data = extra_data
        else:
            self._extra_data = dict()

        self._check()

    def get_extra_data(self, row) -> dict:
        """
        Returns a dictionary with extra data. Values may be callable.
        In this case, the function’s return value is used. The row dict is its
        only argument.
        """
        data = self._extra_data.copy()

        for key in data.keys():
            if callable(data[key]):
                # Call the function, pass row param
                data[key] = data[key](row)

        return data

    def rows(self, exclude=None):
        """
        :param exclude: Iterable of field names to exclude
        :return: A list of row dicts
        """

        rows = self._rows.copy()

        if exclude:
            for record in rows:
                for field in exclude:
                    record.pop(field)

        return rows

    def shall_skip(self, row: dict):
        """
        Indicates whether to skip a row. It return True (i.e. skip) if there’s
        at least one truthy field.
        :param row: A dictionary representing a row
        :return: boolean, skip or not
        """
        values = row.values()
        # When there’s at least one truthy value, don’t skip.
        if any(values):
            return False
        return True

    def load(self, extra_context=None):
        """
        Loads the spreadsheet and map its contents to dicts.
        :return:
        """
        if extra_context is None:
            extra_context = {}

        header_rows = self.get_header_rows()
        fields = self._fields
        sheet = self.__get_sheet()

        if self.has_errors:
            return

        column_map = self.__column_map(sheet)
        self.__check_columns_present(column_map)

        if self.has_errors:
            return

        min_row = header_rows + 1

        for row_index, row_cells in enumerate(sheet.iter_rows(min_row=min_row)):
            row_dict = {}
            error_cache = []
            for name, field in fields.items():
                cell_index = column_map.get(field.source)

                try:
                    cell = row_cells[cell_index]
                    try:
                        row_dict[name] = field(
                            cell.value,
                            language=self.language,
                            extra_context=extra_context,
                        )
                    except ValidationException as error:
                        row_dict[name] = None
                        error_cache.append((str(error), row_index))
                except KeyError:
                    pass
            # Use extra data as a basis
            full_dict = self.get_extra_data(row_dict)
            # And update with “real” data, which takes precedence
            full_dict.update(row_dict)

            if self.shall_skip(full_dict):
                self._add_info("Skipped row", index=row_index)
                continue
            else:
                self._rows.append(full_dict)
                self._add_errors(error_cache)

    def get_sheet_name(self):
        """
        Gets the sheet
        :return: str
        """
        return self.sheet_name

    def get_header_rows(self):
        """
        Gets the number of header rows
        :return: int
        """
        return self.header_rows

    def get_label_row(self):
        """
        Gets the index of the row that contains column labels
        :return: int
        """
        if self.label_row:
            return self.label_row
        return self.get_header_rows() - 1

    @property
    def errors(self):
        return self._errors

    @property
    def has_errors(self):
        return len(self._errors) > 0

    @property
    def infos(self):
        return self._infos

    @property
    def has_infos(self):
        return len(self._infos) > 0

    def __getitem__(self, item):
        if item < 0 or (item > len(self) - 1):
            raise IndexError()

        return self._rows[item]

    def __len__(self):
        return len(self._rows)

    #
    # === Protected ===
    #

    def _build_fields(self) -> dict:
        fields = {}
        for attr, field in self.__class__.__dict__.items():
            if isinstance(field, BaseField):
                fields[attr] = field

        return fields

    def _add_error(self, message, index=None) -> None:
        # Add to row index, if it’s related to a row.
        if isinstance(index, int):
            message = _(
                "sheet.row_info",
                self.language,
                params={
                    "sheet": self.get_sheet_name(),
                    "row": index + 1 + self.get_header_rows(),
                    "message": message,
                },
            )
        self._errors.append(message)

    def _add_errors(self, errors) -> None:
        for error in errors:
            if isinstance(error, tuple):
                error_len = len(error)
                if error_len == 2:
                    self._add_error(error[0], error[1])
                elif error_len == 1:
                    self._add_error(error[0])
                else:
                    raise ValueError("Error tuple too long.")
            else:
                self._add_error(error)

    def _add_info(self, message, index=None) -> None:
        # Add to row index, if it’s related to a row.
        if isinstance(index, int):
            message = _(
                "sheet.row_info",
                self.language,
                params={
                    "sheet": self.get_sheet_name(),
                    "row": index + 1 + self.get_header_rows(),
                    "message": message,
                },
            )
        self._infos.append(message)

    def _check(self) -> None:
        """
        Perform configuration checks. Add your own in subclasses
        :raises ImproperlyConfigured
        """
        if self.get_header_rows() < 1:
            raise ImproperlyConfigured("Sheets must have at least one header row")

        if not self.sheet_name:
            raise ImproperlyConfigured("No sheet name set")

    #
    # === Private ===
    #

    def __column_map(self, sheet) -> dict:
        """
        Takes a sheet and returns a dictionary, that maps column names to column indexes.

        Example:
        ```
        {
            "Artist": 0,
            "Album": 1,
            "Release Date": 2
        }
        ```

        :param sheet: Worksheet from openpyxl
        :return: A dictionary that maps column names to indexes
        """
        column_map = {}
        label_row = self.get_label_row()

        for index, column in enumerate(
            sheet.iter_cols(0, sheet.max_column, min_row=label_row, max_row=label_row)
        ):
            # Retrieve and sanitize the column name from a header cell
            key = str(column[0].value).strip()
            column_map[key] = index

        return column_map

    def __check_columns_present(self, column_map) -> None:
        """
        Checks whether the columns (described by the source attribute of fields) are present.
        Missing columns are added to error list.

        :param column_map: The column map
        """
        fields = self._build_fields()
        used_fields = set([field.source for field in fields.values()])
        all_fields = set(column_map.keys())

        for field in used_fields:
            if field not in all_fields:
                msg = _("sheet.column_missing", self.language, params={"column": field})
                self._add_error(msg)

    def __get_sheet(self):
        wb = load_workbook(filename=self.path, data_only=True)
        sheet_name = self.get_sheet_name()

        try:
            return wb[sheet_name]
        except KeyError:
            msg = _("sheet.sheet_missing", self.language, params={"sheet": sheet_name})
            self._add_error(msg)
